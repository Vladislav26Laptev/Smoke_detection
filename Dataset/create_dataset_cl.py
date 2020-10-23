import os
import cv2
import json
import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(title='Data', index=0)
sheet = wb['Data']
cell = sheet.cell(row=1, column=1)
cell.value = 'Image1_name'
cell = sheet.cell(row=1, column=2)
cell.value = 'Image2_name'
cell = sheet.cell(row=1, column=3)
cell.value = 'Image3_name'
cell = sheet.cell(row=1, column=4)
cell.value = 'Image4_name'
cell = sheet.cell(row=1, column=5)
cell.value = 'Image5_name'
cell = sheet.cell(row=1, column=6)
cell.value = 'Class'
i = 0

# ------------------------------------------------------ No_Smoke ------------------------------------------------------
img_patch = os.listdir('images_cl/')

for img in range(0, len(img_patch), 7):
    image1_name = img_patch[img]
    image2_name = img_patch[img + 1]
    image3_name = img_patch[img + 2]
    image4_name = img_patch[img + 3]
    image5_name = img_patch[img + 4]

    log = [image1_name, image2_name, image3_name, image4_name, image5_name, 0]
    t = 1
    for l in log:
        cell = sheet.cell(row=i + 1, column=t)
        cell.value = l
        t += 1
    i += 1

# --------------------------------------------------------- Smoke ------------------------------------------------------
frames = 7
ann_patch = os.listdir('annotation/')

for ann in ann_patch:
    with open('annotation/' + ann) as f:
        templates = json.load(f)

    width = templates['size']['width']
    height = templates['size']['height']
    TEST_VIDEO_PATH = os.path.join('video', ann[:-5] + '.mp4')
    print(TEST_VIDEO_PATH)
    cap = cv2.VideoCapture(TEST_VIDEO_PATH)
    c = len(templates['frames']) // frames // 2
    for f in range(1, frames-1):
        b = round(len(templates['frames'])/frames*f)
        t = templates['frames'][b]
        box = 0
        for ii in t['figures']:
            if ii['geometryType'] == 'rectangle':
                box += 1
                rectangle = ii['geometry']['points']['exterior']
                x1 = rectangle[0][0]
                y1 = rectangle[0][1]
                x2 = rectangle[1][0]
                if x2 < x1:
                    z = x1
                    x1 = x2
                    x2 = z
                y2 = rectangle[1][1]
                if y2 < y1:
                    z = y1
                    y1 = y2
                    y2 = z

                image1_name = ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) + '_output_img' + '_' + str(1) \
                              + '.png'
                image2_name = ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) + '_output_img' + '_' + str(2) \
                              + '.png'
                image3_name = ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) + '_output_img' + '_' + str(3) \
                              + '.png'
                image4_name = ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) + '_output_img' + '_' + str(4) \
                              + '.png'
                image5_name = ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) + '_output_img' + '_' + str(5) \
                              + '.png'

                log = [image1_name, image2_name, image3_name, image4_name, image5_name, 1]
                t = 1
                for l in log:
                    cell = sheet.cell(row=i + 1, column=t)
                    cell.value = l
                    t += 1
                i += 1

                cap.set(1, b - c * 2)
                _, frame_1 = cap.read()
                frame_1 = frame_1[y1:y2, x1:x2]
                cv2.imwrite('images_cl/' + ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) +
                            '_output_img' + '_' + str(1) + '.png', frame_1)
                cap.set(1, b - c)
                _, frame_2 = cap.read()
                frame_2 = frame_2[y1:y2, x1:x2]
                cv2.imwrite('images_cl/' + ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) +
                            '_output_img' + '_' + str(2) + '.png', frame_2)
                cap.set(1, b)
                _, frame_3 = cap.read()
                frame_3 = frame_3[y1:y2, x1:x2]
                cv2.imwrite('images_cl/' + ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) +
                            '_output_img' + '_' + str(3) + '.png', frame_3)
                cap.set(1, b + c)
                _, frame_4 = cap.read()
                frame_4 = frame_4[y1:y2, x1:x2]
                cv2.imwrite('images_cl/' + ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) +
                            '_output_img' + '_' + str(4) + '.png', frame_4)
                cap.set(1, b + c * 2)
                _, frame_5 = cap.read()
                frame_5 = frame_5[y1:y2, x1:x2]
                cv2.imwrite('images_cl/' + ann[:-5] + '_frame' + str(b).zfill(3) + '_box' + str(box) +
                            '_output_img' + '_' + str(5) + '.png', frame_5)

    cap.release()
    print('video_' + ann[0:7] + '_check!')

wb.save('dataset_label_cl.xlsx')
